"""
validate_excel_vs_warehouse.py — Compara el Excel con el warehouse
================================================================
Verifica diferencias entre:
  - Empresas en Excel vs warehouse
  - Industrias en Excel vs warehouse
  - Tipos de empresa (DGI vs VACA LECHERA)
  - Datos financieros
"""

import sqlite3
import pandas as pd
import openpyxl
from pathlib import Path

DB_PATH = "output/warehouse.db"
EXCEL_PATH = "g:/Code/PlanillaCursoDividendos.xlsx"

def get_excel_empresas():
    """Extrae la lista de empresas del Excel"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # Hoja "Empresas Power BI"
    ws = wb["Empresas Power BI"]
    empresas_excel = []
    
    # Saltar encabezados (fila 1)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Columna A = Empresas
            empresas_excel.append({
                'empresa': row[0],
                'industria': row[1],
                'tipo': row[2]
            })
    
    wb.close()
    return empresas_excel

def get_warehouse_empresas():
    """Obtiene las empresas del warehouse"""
    con = sqlite3.connect(DB_PATH)
    
    # Obtener tickers únicos y su industria
    query = """
    SELECT DISTINCT ticker, industry 
    FROM normalized_financials 
    ORDER BY ticker
    """
    df = pd.read_sql(query, con)
    con.close()
    
    return df.to_dict('records')

def get_warehouse_stats():
    """Estadísticas del warehouse"""
    con = sqlite3.connect(DB_PATH)
    stats = {}
    
    # Total de filas
    stats['total_rows'] = pd.read_sql("SELECT COUNT(*) as c FROM normalized_financials", con).iloc[0]['c']
    
    # Tickers únicos
    stats['tickers'] = pd.read_sql("SELECT COUNT(DISTINCT ticker) as c FROM normalized_financials", con).iloc[0]['c']
    
    # Por industria
    stats['by_industry'] = pd.read_sql("""
        SELECT industry, COUNT(DISTINCT ticker) as tickers, COUNT(*) as rows
        FROM normalized_financials 
        GROUP BY industry
    """, con).to_dict('records')
    
    # Años disponibles
    stats['years'] = sorted(pd.read_sql("SELECT DISTINCT year FROM normalized_financials ORDER BY year", con)['year'].tolist())
    
    # Tickers por año
    stats['tickers_by_year'] = pd.read_sql("""
        SELECT year, COUNT(DISTINCT ticker) as tickers
        FROM normalized_financials 
        GROUP BY year 
        ORDER BY year
    """, con).to_dict('records')
    
    con.close()
    return stats

def compare_with_excel():
    """Comparación principal"""
    print("=" * 70)
    print("COMPARACIÓN EXCEL vs WAREHOUSE")
    print("=" * 70)
    
    # 1. Obtener datos del Excel
    print("\n1. EXCEL - Hoja 'Empresas Power BI':")
    print("-" * 40)
    empresas_excel = get_excel_empresas()
    
    print(f"   Total empresas en Excel: {len(empresas_excel)}")
    
    # Clasificar por tipo
    dgi = [e for e in empresas_excel if e['tipo'] == 'DGI']
    vaca = [e for e in empresas_excel if e['tipo'] == 'VACA LECHERA']
    print(f"   DGI: {len(dgi)}")
    print(f"   VACA LECHERA: {len(vaca)}")
    
    # Mostrar industrias
    industrias = {}
    for e in empresas_excel:
        ind = e['industria']
        if ind:
            industrias[ind] = industrias.get(ind, 0) + 1
    print(f"\n   Industrias en Excel:")
    for ind, count in sorted(industrias.items()):
        print(f"      {ind}: {count}")
    
    # 2. Obtener datos del warehouse
    print("\n2. WAREHOUSE (output/warehouse.db):")
    print("-" * 40)
    warehouse_empresas = get_warehouse_empresas()
    stats = get_warehouse_stats()
    
    print(f"   Total filas: {stats['total_rows']:,}")
    print(f"   Tickers únicos: {stats['tickers']}")
    print(f"   Años: {stats['years'][0]} - {stats['years'][-1]} ({len(stats['years'])} años)")
    
    print(f"\n   Por industria:")
    for row in stats['by_industry']:
        print(f"      {row['industry']}: {row['tickers']} tickers, {row['rows']:,} rows")
    
    # 3. Comparar tickers
    print("\n3. DIFERENCIAS DE TICKERS:")
    print("-" * 40)
    
    tickers_excel = {e['empresa'].upper().replace(' ', '').replace('-', ''): e for e in empresas_excel}
    tickers_warehouse = {e['ticker'].upper().replace('.', '').replace('-', ''): e for e in warehouse_empresas}
    
    # Empresas en Excel pero NO en warehouse
    missing_in_warehouse = []
    for te, data in tickers_excel.items():
        # Normalizar nombres
        found = False
        for tw in tickers_warehouse:
            if te in tw or tw in te:
                found = True
                break
        if not found:
            missing_in_warehouse.append(data['empresa'])
    
    if missing_in_warehouse:
        print(f"\n   En Excel pero NO en warehouse ({len(missing_in_warehouse)}):")
        for e in sorted(missing_in_warehouse):
            print(f"      - {e}")
    else:
        print("   ✓ Todas las empresas del Excel están en el warehouse")
    
    # Empresas en warehouse pero NO en Excel
    extra_in_warehouse = []
    for tw, data in tickers_warehouse.items():
        found = False
        for te in tickers_excel:
            if te in tw or tw in te:
                found = True
                break
        if not found:
            extra_in_warehouse.append(data['ticker'])
    
    if extra_in_warehouse:
        print(f"\n   En warehouse pero NO en Excel ({len(extra_in_warehouse)}):")
        for e in sorted(extra_in_warehouse):
            print(f"      + {e}")
    
    # 4. Comparar industrias
    print("\n4. ANÁLISIS DE INDUSTRIAS:")
    print("-" * 40)
    
    # Industrias del Excel
    excel_industries = set()
    for e in empresas_excel:
        if e['industria']:
            excel_industries.add(e['industria'].upper())
    
    # Industrias del warehouse
    wh_industries = set()
    for e in warehouse_empresas:
        if e['industry']:
            wh_industries.add(e['industry'].upper())
    
    print(f"\n   Industrias en Excel: {len(excel_industries)}")
    print(f"   Industrias en warehouse: {len(wh_industries)}")
    
    missing_ind = excel_industries - wh_industries
    if missing_ind:
        print(f"\n   Industrias en Excel pero NO en warehouse:")
        for ind in sorted(missing_ind):
            print(f"      - {ind}")
    
    extra_ind = wh_industries - excel_industries
    if extra_ind:
        print(f"\n   Industrias en warehouse pero NO en Excel:")
        for ind in sorted(extra_ind):
            print(f"      + {ind}")
    
    # 5. Verificar hojas individuales del Excel vs datos warehouse
    print("\n5. VERIFICACIÓN DE DATOS POR EMPRESA:")
    print("-" * 40)
    
    # Leer algunas hojas específicas
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet_names = wb.sheetnames
    
    # Hojas de empresas (excluyendo hojas de documentación)
    doc_sheets = ['Responsables empresas', 'Cuidados a tener en la planilla', 
                  'Actualización de POWER BI', 'Como llenar dividendos', 
                  'Empresas Power BI', 'INDICE']
    empresa_sheets = [s for s in sheet_names if s not in doc_sheets]
    
    print(f"   Hojas de empresas en Excel: {len(empresa_sheets)}")
    
    # Verificar si las hojas tienen datos
    sheets_with_data = []
    sheets_empty = []
    
    for sheet in empresa_sheets[:10]:  # Solo las primeras 10 para no saturar
        ws = wb[sheet]
        # Verificar si tiene datos (filas con valores)
        has_data = False
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if any(cell is not None for cell in row):
                has_data = True
                break
        
        if has_data:
            sheets_with_data.append(sheet)
        else:
            sheets_empty.append(sheet)
    
    print(f"\n   Hojas con datos (muestra): {len(sheets_with_data)}")
    for s in sheets_with_data[:5]:
        print(f"      ✓ {s}")
    
    if sheets_empty:
        print(f"\n   Hojas vacías o sin estructura: {len(sheets_empty)}")
        for s in sheets_empty[:5]:
            print(f"      ✗ {s}")
    
    wb.close()
    
    # 6. Resumen final
    print("\n" + "=" * 70)
    print("RESUMEN DE DIFERENCIAS")
    print("=" * 70)
    
    print(f"""
    EXCEL:
    - Empresas esperadas: {len(empresas_excel)}
    - Industrias: {len(industrias)}
    - DGI: {len(dgi)}, VACA LECHERA: {len(vaca)}
    
    WAREHOUSE:
    - Tickers: {stats['tickers']}
    - Industrias: {len(wh_industries)}
    - Filas totales: {stats['total_rows']:,}
    - Años: {len(stats['years'])}
    
    DIFERENCIAS:
    - Faltan en warehouse: {len(missing_in_warehouse)} empresas
    - Extra en warehouse: {len(extra_in_warehouse)} empresas
    """)

if __name__ == "__main__":
    compare_with_excel()